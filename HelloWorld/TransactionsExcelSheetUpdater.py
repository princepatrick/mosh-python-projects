import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for i in range(2, sheet.max_row+1):
        current_cost = sheet.cell(i, 3)
        revised_cost = current_cost.value * 0.9
        revised_cost_cell = sheet.cell(i, 4)
        revised_cost_cell.value = revised_cost

    revised_cost_cell_header = sheet.cell(1, 4)
    revised_cost_cell_header.value = "revised price"

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")

    wb.save(filename)

